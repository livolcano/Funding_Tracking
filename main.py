from __future__ import annotations

import argparse
import csv
import logging
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from time import sleep
from typing import Any
from urllib.parse import urljoin

import httpx
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup, Tag


KAKEN_BASE_URL = "https://kaken.nii.ac.jp"
DEFAULT_INPUT_EXCEL_DIR = Path("input") / "researcher_excels"
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/126.0.0.0 Safari/537.36"
    )
}
FUNDED_STATUS_KEYWORDS = ("adopted", "granted", "completed", "ongoing", "accepted")
RETRY_ATTEMPTS = 3
RETRY_BASE_DELAY_SECONDS = 1.0
DEFAULT_PROJECT_LIST_EXCEL = Path("output") / "kaken_project_list.xlsx"


@dataclass
class SearchCandidate:
    project_id: str
    project_title: str
    detail_url: str
    pi_name: str
    institution: str
    project_period_text: str
    project_status: str


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().lower()


def parse_date_range(text: str) -> tuple[str | None, str | None]:
    match = re.search(r"(\d{4}-\d{2}-\d{2})\s*[–-]\s*(\d{4}-\d{2}-\d{2})", text)
    if not match:
        return None, None
    return match.group(1), match.group(2)


def parse_budget_total_jpy(budget_text: str) -> int | None:
    match = re.search(r"[¥￥]\s*([0-9][0-9,]*)", budget_text)
    if not match:
        return None
    return int(match.group(1).replace(",", ""))


def extract_local_name_variant(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    non_ascii_chunks = re.findall(r"[^\x00-\x7F]+(?:\s+[^\x00-\x7F]+)*", text)
    if not non_ascii_chunks:
        return ""
    return normalize_text(max(non_ascii_chunks, key=len))


def is_funded_status(status: str) -> bool:
    lowered = status.lower()
    return any(keyword in lowered for keyword in FUNDED_STATUS_KEYWORDS)


def is_completed_status(status: str) -> bool:
    return "completed" in status.lower()


def extract_year(date_text: str | None) -> int | None:
    if not date_text:
        return None
    match = re.match(r"(\d{4})-\d{2}-\d{2}", date_text)
    if not match:
        return None
    return int(match.group(1))


def table_rows(summary_table: Tag | None) -> dict[str, Tag]:
    if summary_table is None:
        return {}
    rows: dict[str, Tag] = {}
    for row in summary_table.select("tr"):
        th = row.select_one("th")
        td = row.select_one("td")
        if th is None or td is None:
            continue
        rows[normalize_key(th.get_text(" ", strip=True))] = td
    return rows


def request_get_with_retry(
    client: httpx.Client,
    url: str,
    *,
    params: dict[str, str] | None = None,
    attempts: int = RETRY_ATTEMPTS,
    base_delay_seconds: float = RETRY_BASE_DELAY_SECONDS,
) -> httpx.Response:
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        try:
            response = client.get(url, params=params)
            response.raise_for_status()
            return response
        except (httpx.TimeoutException, httpx.NetworkError, httpx.RemoteProtocolError) as exc:
            last_error = exc
            if attempt >= attempts:
                break
            backoff_seconds = base_delay_seconds * (2 ** (attempt - 1))
            logging.warning(
                "Request failed for %s (attempt %d/%d): %s. Retrying in %.1fs",
                url,
                attempt,
                attempts,
                exc,
                backoff_seconds,
            )
            sleep(backoff_seconds)
    raise RuntimeError(f"Request failed after {attempts} attempts for URL: {url}") from last_error


def fetch_search_candidates(
    client: httpx.Client,
    query: str,
    page_size: int,
    max_pages: int,
) -> list[SearchCandidate]:
    all_candidates: list[SearchCandidate] = []
    for page_index in range(max_pages):
        start = page_index * page_size + 1
        params = {"kw": query, "rw": str(page_size), "od": "2", "st": str(start)}
        response = request_get_with_retry(client, urljoin(KAKEN_BASE_URL, "/en/search/"), params=params)
        soup = BeautifulSoup(response.text, "html.parser")
        candidates = parse_search_candidates(soup)
        if not candidates:
            break
        all_candidates.extend(candidates)
        if len(candidates) < page_size:
            break
    return all_candidates


def fetch_keyword_project_list(
    client: httpx.Client,
    keyword_query: str,
    page_size: int,
    max_pages: int,
    only_not_completed: bool,
    started_after_year: int | None,
    top_n: int,
) -> list[dict[str, Any]]:
    candidates = fetch_search_candidates(client, keyword_query, page_size=page_size, max_pages=max_pages)
    filtered: list[dict[str, Any]] = []
    seen_project_ids: set[str] = set()
    for candidate in candidates:
        if candidate.project_id in seen_project_ids:
            continue
        seen_project_ids.add(candidate.project_id)

        if candidate.project_status and not is_funded_status(candidate.project_status):
            continue
        if only_not_completed and is_completed_status(candidate.project_status):
            continue

        project_start, project_end = parse_date_range(candidate.project_period_text)
        project_start_year = extract_year(project_start)
        if started_after_year is not None:
            if project_start_year is None or project_start_year <= started_after_year:
                continue

        filtered.append(
            {
                "search_keyword": keyword_query,
                "project_id": candidate.project_id,
                "project_title": candidate.project_title,
                "project_status": candidate.project_status,
                "project_start": project_start or "",
                "project_end": project_end or "",
                "project_pi_name": candidate.pi_name,
                "project_institution": candidate.institution,
                "project_period_text": candidate.project_period_text,
                "detail_url": candidate.detail_url,
            }
        )

    filtered.sort(
        key=lambda item: (
            item["project_start"] or "0000-00-00",
            item["project_id"],
        ),
        reverse=True,
    )
    return filtered[:top_n]


def parse_search_candidates(soup: BeautifulSoup) -> list[SearchCandidate]:
    results: list[SearchCandidate] = []
    for item in soup.select("div.listContainer > ul > li div.listitem"):
        title_anchor = item.select_one("h3 .title a.link-page[href*='/en/grant/']")
        if title_anchor is None:
            continue
        href = title_anchor.get("href", "").strip()
        if not href:
            continue
        project_id = href.strip("/").split("/")[-1]
        detail_url = urljoin(KAKEN_BASE_URL, href)
        project_title = normalize_text(title_anchor.get_text(" ", strip=True))

        rows = table_rows(item.select_one("table.summary-table"))
        pi_td = rows.get("principal investigator")
        pi_name = ""
        if pi_td is not None:
            pi_anchor = pi_td.select_one("span a")
            if pi_anchor is not None:
                pi_name = normalize_text(pi_anchor.get_text(" ", strip=True))
            else:
                pi_name = normalize_text(pi_td.get_text(" ", strip=True))

        institution_td = rows.get("research institution")
        institution = ""
        if institution_td is not None:
            institution = normalize_text(institution_td.get_text(" ", strip=True))

        period_td = rows.get("project period (fy)")
        project_period_text = ""
        project_status = ""
        if period_td is not None:
            project_period_text = normalize_text(period_td.get_text(" ", strip=True))
            status_tag = period_td.select_one("span.pstatus")
            if status_tag is not None:
                project_status = normalize_text(status_tag.get_text(" ", strip=True))

        results.append(
            SearchCandidate(
                project_id=project_id,
                project_title=project_title,
                detail_url=detail_url,
                pi_name=pi_name,
                institution=institution,
                project_period_text=project_period_text,
                project_status=project_status,
            )
        )
    return results


def select_recent_funded_candidate(
    candidates: list[SearchCandidate],
    researcher_name: str,
) -> SearchCandidate | None:
    if not candidates:
        return None
    normalized_name = normalize_key(researcher_name)
    matched_by_name = [
        c for c in candidates if normalized_name and normalized_name in normalize_key(c.pi_name)
    ]
    pool = matched_by_name or candidates
    funded = [c for c in pool if c.project_status and is_funded_status(c.project_status)]
    if funded:
        return funded[0]
    return pool[0]


def parse_detail_project_info(client: httpx.Client, detail_url: str) -> dict[str, Any]:
    response = request_get_with_retry(client, detail_url)
    soup = BeautifulSoup(response.text, "html.parser")

    page_title = soup.select_one("div.page-title h1")
    project_title = ""
    if page_title is not None:
        project_title = normalize_text(page_title.get_text(" ", strip=True))

    rows = table_rows(soup.select_one("table.summary-table"))

    keyword_td = rows.get("keywords") or rows.get("keyword")
    keywords = ""
    if keyword_td is not None:
        anchors = [normalize_text(a.get_text(" ", strip=True)) for a in keyword_td.select("a")]
        if anchors:
            keywords = "; ".join([text for text in anchors if text])
        else:
            keywords = normalize_text(keyword_td.get_text(" ", strip=True))

    budget_td = None
    for key, cell in rows.items():
        if key.startswith("budget amount"):
            budget_td = cell
            break
    budget_total = ""
    if budget_td is not None:
        h5 = budget_td.select_one("h5")
        if h5 is not None:
            budget_total = normalize_text(h5.get_text(" ", strip=True))
        else:
            budget_total = normalize_text(budget_td.get_text(" ", strip=True))

    period_td = rows.get("project period (fy)")
    period_text = normalize_text(period_td.get_text(" ", strip=True)) if period_td is not None else ""
    project_start, project_end = parse_date_range(period_text)

    pi_td = rows.get("principal investigator")
    project_pi_name_local = ""
    if pi_td is not None:
        pi_anchor = pi_td.select_one("a[href*='nrid.nii.ac.jp']")
        if pi_anchor is not None:
            profile_url = str(pi_anchor.get("href", "")).strip()
            profile_url = profile_url.replace("/en/nrid/", "/nrid/")
            if profile_url:
                try:
                    profile_resp = request_get_with_retry(client, profile_url)
                    profile_soup = BeautifulSoup(profile_resp.text, "html.parser")
                    profile_h1 = profile_soup.select_one("h1")
                    if profile_h1 is not None:
                        project_pi_name_local = extract_local_name_variant(
                            profile_h1.get_text(" ", strip=True)
                        )
                except httpx.HTTPError:
                    project_pi_name_local = ""

    return {
        "project_title": project_title,
        "keywords": keywords,
        "budget_total": budget_total,
        "project_start": project_start,
        "project_end": project_end,
        "project_pi_name_local": project_pi_name_local,
    }


def fetch_latest_jpy_to_usd_rate(client: httpx.Client) -> float:
    response = request_get_with_retry(client, "https://open.er-api.com/v6/latest/JPY")
    payload = response.json()
    rates = payload.get("rates", {})
    rate = rates.get("USD")
    if rate is None:
        raise ValueError("USD exchange rate not found in API response.")
    return float(rate)


def save_to_sqlite(records: list[dict[str, Any]], sqlite_path: Path) -> None:
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(sqlite_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS kaken_latest_projects (
                researcher_name TEXT NOT NULL,
                affiliation TEXT,
                query TEXT NOT NULL,
                project_id TEXT,
                project_title TEXT,
                keywords TEXT,
                budget_total TEXT,
                budget_total_usd TEXT,
                project_start TEXT,
                project_end TEXT,
                project_status TEXT,
                project_pi_name TEXT,
                project_pi_name_local TEXT,
                project_institution TEXT,
                detail_url TEXT,
                fetched_at_utc TEXT NOT NULL,
                PRIMARY KEY (researcher_name, affiliation)
            )
            """
        )
        existing_columns = {
            row[1] for row in conn.execute("PRAGMA table_info(kaken_latest_projects)").fetchall()
        }
        if "budget_total_usd" not in existing_columns:
            conn.execute("ALTER TABLE kaken_latest_projects ADD COLUMN budget_total_usd TEXT")
        if "project_pi_name_local" not in existing_columns:
            conn.execute("ALTER TABLE kaken_latest_projects ADD COLUMN project_pi_name_local TEXT")
        conn.executemany(
            """
            INSERT INTO kaken_latest_projects (
                researcher_name, affiliation, query, project_id, project_title, keywords,
                budget_total, budget_total_usd, project_start, project_end, project_status,
                project_pi_name, project_pi_name_local, project_institution, detail_url, fetched_at_utc
            ) VALUES (
                :researcher_name, :affiliation, :query, :project_id, :project_title, :keywords,
                :budget_total, :budget_total_usd, :project_start, :project_end, :project_status,
                :project_pi_name, :project_pi_name_local, :project_institution, :detail_url, :fetched_at_utc
            )
            ON CONFLICT(researcher_name, affiliation) DO UPDATE SET
                query = excluded.query,
                project_id = excluded.project_id,
                project_title = excluded.project_title,
                keywords = excluded.keywords,
                budget_total = excluded.budget_total,
                budget_total_usd = excluded.budget_total_usd,
                project_start = excluded.project_start,
                project_end = excluded.project_end,
                project_status = excluded.project_status,
                project_pi_name = excluded.project_pi_name,
                project_pi_name_local = excluded.project_pi_name_local,
                project_institution = excluded.project_institution,
                detail_url = excluded.detail_url,
                fetched_at_utc = excluded.fetched_at_utc
            """,
            records,
        )
        conn.commit()


def build_query(name: str, affiliation: str) -> str:
    parts = [normalize_text(part) for part in [affiliation, name] if normalize_text(part)]
    return " ".join(parts)


def parse_sheet_selector(raw_sheet_name: str | int | None) -> str | int:
    if isinstance(raw_sheet_name, int):
        return raw_sheet_name
    if raw_sheet_name is None:
        return 0
    parsed = normalize_text(str(raw_sheet_name))
    if parsed.isdigit():
        return int(parsed)
    return parsed


def resolve_input_excel_path(input_path: Path | None, input_dir: Path) -> Path:
    excel_suffixes = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if input_path is not None:
        if input_path.is_dir():
            candidates = sorted(
                path for path in input_path.iterdir() if path.is_file() and path.suffix.lower() in excel_suffixes
            )
            if not candidates:
                raise FileNotFoundError(f"No Excel files found in: {input_path}")
            return candidates[0]
        return input_path

    candidates = sorted(
        path for path in input_dir.iterdir() if path.is_file() and path.suffix.lower() in excel_suffixes
    )
    if not candidates:
        raise FileNotFoundError(f"No Excel files found in default input directory: {input_dir}")
    return candidates[0]


def read_researchers_from_excel(
    input_excel: Path,
    sheet_name: str | int | None,
    name_column: str,
    affiliation_column: str,
    max_researchers: int | None,
) -> list[dict[str, str]]:
    wb = load_workbook(input_excel, read_only=True, data_only=True)
    selector = parse_sheet_selector(sheet_name)
    if isinstance(selector, int):
        if selector < 0 or selector >= len(wb.worksheets):
            raise ValueError(f"sheet index out of range: {selector}")
        ws = wb.worksheets[selector]
    else:
        if selector not in wb.sheetnames:
            raise ValueError(f"sheet not found: {selector}")
        ws = wb[selector]

    rows = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows)
    except StopIteration as exc:
        raise ValueError("Excel sheet is empty.") from exc
    headers = [normalize_text(str(cell)) if cell is not None else "" for cell in headers_raw]
    if name_column not in headers:
        raise ValueError(f"Excel missing required column: {name_column}")
    if affiliation_column not in headers:
        raise ValueError(f"Excel missing required column: {affiliation_column}")

    name_idx = headers.index(name_column)
    affiliation_idx = headers.index(affiliation_column)
    records: list[dict[str, str]] = []
    for row in rows:
        raw_name = row[name_idx] if name_idx < len(row) else None
        raw_affiliation = row[affiliation_idx] if affiliation_idx < len(row) else None
        researcher_name = normalize_text("" if raw_name is None else str(raw_name))
        affiliation = normalize_text("" if raw_affiliation is None else str(raw_affiliation))
        if not researcher_name:
            continue
        records.append({"researcher_name": researcher_name, "affiliation": affiliation})
        if max_researchers is not None and len(records) >= max_researchers:
            break
    return records


def write_csv(records: list[dict[str, Any]], output_csv: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(records[0].keys())
    with output_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def write_project_list_excel(records: list[dict[str, Any]], output_excel: Path) -> None:
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "project_list"

    if not records:
        ws.append(
            [
                "search_keyword",
                "project_id",
                "project_title",
                "project_status",
                "project_start",
                "project_end",
                "project_pi_name",
                "project_institution",
                "project_period_text",
                "detail_url",
            ]
        )
    else:
        headers = list(records[0].keys())
        ws.append(headers)
        for record in records:
            ws.append([record.get(header, "") for header in headers])
    wb.save(output_excel)


def process_keyword_queries(
    keyword_queries: list[str],
    page_size: int,
    max_pages: int,
    top_n_per_keyword: int,
    only_not_completed: bool,
    started_after_year: int | None,
    verify_ssl: bool,
    ca_bundle: str | None,
) -> list[dict[str, Any]]:
    verify_config: bool | str = verify_ssl
    if ca_bundle:
        verify_config = ca_bundle
    all_records: list[dict[str, Any]] = []

    with httpx.Client(
        headers=DEFAULT_HEADERS,
        timeout=30.0,
        follow_redirects=True,
        verify=verify_config,
    ) as client:
        for keyword_query in keyword_queries:
            cleaned_keyword = normalize_text(keyword_query)
            if not cleaned_keyword:
                continue
            logging.info("Searching funded projects by keyword [%s]", cleaned_keyword)
            records = fetch_keyword_project_list(
                client=client,
                keyword_query=cleaned_keyword,
                page_size=page_size,
                max_pages=max_pages,
                only_not_completed=only_not_completed,
                started_after_year=started_after_year,
                top_n=top_n_per_keyword,
            )
            all_records.extend(records)
    return all_records


def process_researchers(
    input_excel: Path,
    sheet_name: str | int | None,
    name_column: str,
    affiliation_column: str,
    page_size: int,
    max_pages: int,
    max_researchers: int | None,
    delay_seconds: float,
    verify_ssl: bool,
    ca_bundle: str | None,
) -> list[dict[str, Any]]:
    source_rows = read_researchers_from_excel(
        input_excel=input_excel,
        sheet_name=sheet_name,
        name_column=name_column,
        affiliation_column=affiliation_column,
        max_researchers=max_researchers,
    )

    records: list[dict[str, Any]] = []
    verify_config: bool | str = verify_ssl
    if ca_bundle:
        verify_config = ca_bundle
    with httpx.Client(
        headers=DEFAULT_HEADERS,
        timeout=30.0,
        follow_redirects=True,
        verify=verify_config,
    ) as client:
        jpy_to_usd_rate = fetch_latest_jpy_to_usd_rate(client)
        logging.info("Using JPY->USD rate: %.6f", jpy_to_usd_rate)
        for index, row in enumerate(source_rows):
            researcher_name = row["researcher_name"]
            affiliation = row["affiliation"]
            query = build_query(researcher_name, affiliation)
            if not query:
                raise ValueError(
                    f"Cannot build query at row index {index}; check name/affiliation values."
                )

            logging.info("Searching KAKEN for [%s]", query)
            candidates = fetch_search_candidates(client, query, page_size=page_size, max_pages=max_pages)
            selected = select_recent_funded_candidate(candidates, researcher_name)
            fetched_at = datetime.now(timezone.utc).isoformat()

            if selected is None:
                records.append(
                    {
                        "researcher_name": researcher_name,
                        "affiliation": affiliation,
                        "query": query,
                        "project_id": "",
                        "project_title": "",
                        "keywords": "",
                        "budget_total": "",
                        "budget_total_usd": "",
                        "project_start": "",
                        "project_end": "",
                        "project_status": "",
                        "project_pi_name": "",
                        "project_pi_name_local": "",
                        "project_institution": "",
                        "detail_url": "",
                        "fetched_at_utc": fetched_at,
                    }
                )
                logging.warning("No project found for [%s]", query)
                sleep(delay_seconds)
                continue

            detail = parse_detail_project_info(client, selected.detail_url)
            budget_total_jpy = parse_budget_total_jpy(detail["budget_total"])
            budget_total_usd = ""
            if budget_total_jpy is not None:
                budget_total_usd = f"{budget_total_jpy * jpy_to_usd_rate:.2f}"
            records.append(
                {
                    "researcher_name": researcher_name,
                    "affiliation": affiliation,
                    "query": query,
                    "project_id": selected.project_id,
                    "project_title": detail["project_title"] or selected.project_title,
                    "keywords": detail["keywords"],
                    "budget_total": detail["budget_total"],
                    "budget_total_usd": budget_total_usd,
                    "project_start": detail["project_start"] or "",
                    "project_end": detail["project_end"] or "",
                    "project_status": selected.project_status,
                    "project_pi_name": selected.pi_name,
                    "project_pi_name_local": detail["project_pi_name_local"],
                    "project_institution": selected.institution,
                    "detail_url": selected.detail_url,
                    "fetched_at_utc": fetched_at,
                }
            )
            sleep(delay_seconds)
    return records


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Search KAKEN with researcher names/affiliations from a local Excel file and "
            "extract each researcher's latest funded project details."
        )
    )
    parser.add_argument(
        "--input-excel",
        default=None,
        help="Path to input Excel file. If omitted, the first Excel file in the input directory is used.",
    )
    parser.add_argument(
        "--input-dir",
        default=str(DEFAULT_INPUT_EXCEL_DIR),
        help="Directory containing Excel files to search.",
    )
    parser.add_argument("--sheet-name", default=0, help="Excel sheet name or zero-based sheet index.")
    parser.add_argument(
        "--name-column",
        default="researcher_name",
        help="Column name for researcher name in Excel.",
    )
    parser.add_argument(
        "--affiliation-column",
        default="affiliation",
        help="Column name for school/university affiliation in Excel.",
    )
    parser.add_argument(
        "--output-csv",
        default="output\\kaken_latest_projects.csv",
        help="Path to output CSV file.",
    )
    parser.add_argument(
        "--output-sqlite",
        default="output\\kaken_latest_projects.db",
        help="Path to output SQLite database file.",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=100,
        choices=[20, 50, 100, 200, 500],
        help="KAKEN results per page.",
    )
    parser.add_argument("--max-pages", type=int, default=3, help="Max pages to scan per researcher query.")
    parser.add_argument("--max-researchers", type=int, default=None, help="Optional cap for trial runs.")
    parser.add_argument(
        "--delay-seconds",
        type=float,
        default=1.0,
        help="Delay between each researcher query to avoid overloading the site.",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging level.",
    )
    parser.add_argument(
        "--ca-bundle",
        default=None,
        help="Optional custom CA bundle path for TLS verification.",
    )
    parser.add_argument(
        "--insecure-skip-verify",
        action="store_true",
        help="Disable TLS certificate verification (use only in trusted environments).",
    )
    parser.add_argument(
        "--keyword-query",
        action="append",
        default=[],
        help="Keyword query for project-list mode. Repeat this option for multiple keywords.",
    )
    parser.add_argument(
        "--project-list-excel",
        default=str(DEFAULT_PROJECT_LIST_EXCEL),
        help="Output Excel path for keyword project-list mode.",
    )
    parser.add_argument(
        "--project-only-not-completed",
        action="store_true",
        help="In keyword project-list mode, keep only projects that are not completed.",
    )
    parser.add_argument(
        "--project-started-after-year",
        type=int,
        default=None,
        help="In keyword project-list mode, keep only projects with start year greater than this year.",
    )
    parser.add_argument(
        "--project-top-n-per-keyword",
        type=int,
        default=50,
        help="In keyword project-list mode, max number of projects to keep per keyword after filtering.",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level), format="%(levelname)s %(message)s")
    if args.insecure_skip_verify:
        logging.warning("TLS certificate verification is disabled for this run.")

    if args.keyword_query:
        project_records = process_keyword_queries(
            keyword_queries=args.keyword_query,
            page_size=args.page_size,
            max_pages=args.max_pages,
            top_n_per_keyword=args.project_top_n_per_keyword,
            only_not_completed=args.project_only_not_completed,
            started_after_year=args.project_started_after_year,
            verify_ssl=not args.insecure_skip_verify,
            ca_bundle=args.ca_bundle,
        )
        project_list_excel = Path(args.project_list_excel)
        write_project_list_excel(project_records, project_list_excel)
        logging.info("Project list Excel saved: %s", project_list_excel)
        logging.info("Done. Total projects listed: %d", len(project_records))
        return

    input_dir = Path(args.input_dir)
    if args.input_excel is not None:
        input_excel = resolve_input_excel_path(Path(args.input_excel), input_dir)
    else:
        if not input_dir.exists():
            raise FileNotFoundError(f"Input Excel directory not found: {input_dir}")
        input_excel = resolve_input_excel_path(None, input_dir)

    if not input_excel.exists():
        raise FileNotFoundError(f"Input Excel file not found: {input_excel}")
    logging.info("Using input Excel: %s", input_excel)

    records = process_researchers(
        input_excel=input_excel,
        sheet_name=args.sheet_name,
        name_column=args.name_column,
        affiliation_column=args.affiliation_column,
        page_size=args.page_size,
        max_pages=args.max_pages,
        max_researchers=args.max_researchers,
        delay_seconds=args.delay_seconds,
        verify_ssl=not args.insecure_skip_verify,
        ca_bundle=args.ca_bundle,
    )
    if not records:
        raise RuntimeError("No records were produced.")

    output_csv = Path(args.output_csv)
    write_csv(records, output_csv)
    logging.info("CSV saved: %s", output_csv)

    output_sqlite = Path(args.output_sqlite)
    save_to_sqlite(records, output_sqlite)
    logging.info("SQLite saved: %s", output_sqlite)
    logging.info("Done. Total researchers processed: %d", len(records))


if __name__ == "__main__":
    main()
